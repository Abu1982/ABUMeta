"""把 Excel/CSV 外贸询盘映射到 3-Agent 风险样本。"""

from __future__ import annotations

import csv
from dataclasses import dataclass
import json
from pathlib import Path
import re
from typing import Any, Iterable, Optional, cast

from src.social.trade_warning import (
    AgentRosterEntry,
    TradeRiskSample,
    TradeRiskSampleSet,
)


FREE_EMAIL_DOMAINS = {
    "gmail.com",
    "outlook.com",
    "hotmail.com",
    "yahoo.com",
    "yahoo.co.jp",
    "qq.com",
    "163.com",
    "126.com",
}

HEADER_ALIASES = {
    "sample_id": ["sample_id", "inquiry_id", "询盘编号", "线索编号", "id"],
    "scenario_name": ["scenario_name", "场景标题", "标题", "subject", "询盘标题"],
    "counterparty_name": [
        "counterparty_name",
        "buyer_name",
        "客户名称",
        "公司名称",
        "客户",
    ],
    "destination_country": [
        "destination_country",
        "country",
        "destination",
        "location",
        "国家",
        "目的国",
    ],
    "product_name": ["product_name", "product", "title", "产品", "品名"],
    "quantity": ["quantity", "qty", "数量", "采购数量"],
    "product_category": ["product_category", "product_type", "产品分类", "类目"],
    "trade_value_usd": [
        "trade_value_usd",
        "order_value_usd",
        "amount_usd",
        "金额usd",
        "订单金额",
    ],
    "payment_terms": [
        "payment_terms",
        "payment_method",
        "paymentmode",
        "付款条件",
        "付款方式",
        "账期条款",
    ],
    "requirement": ["requirement", "需求周期", "交期要求", "requirementtime"],
    "frequency": ["frequency", "采购频次", "下单频率"],
    "payment_days": ["payment_days", "账期天数", "账期"],
    "deposit_ratio": ["deposit_ratio", "deposit", "首付款比例", "预付款比例"],
    "incoterm": ["incoterm", "trade_term", "贸易术语", "交货条款"],
    "requested_discount_ratio": [
        "requested_discount_ratio",
        "discount",
        "折扣比例",
        "议价幅度",
    ],
    "account_age_days": [
        "account_age_days",
        "客户年限天数",
        "主体建立天数",
        "公司成立天数",
    ],
    "email": ["email", "邮箱", "联系人邮箱"],
    "uses_free_email": ["uses_free_email", "免费邮箱"],
    "asks_undervalue_invoice": ["asks_undervalue_invoice", "低报发票", "是否要求低报"],
    "asks_split_payment_to_third_party": [
        "asks_split_payment_to_third_party",
        "第三方收款",
        "是否第三方收款",
    ],
    "asks_release_bl_before_payment": [
        "asks_release_bl_before_payment",
        "先放单后付款",
        "提单控制权",
    ],
    "hs_code_sensitive": ["hs_code_sensitive", "敏感hs", "hs敏感", "敏感编码"],
    "sanction_watch_hit": ["sanction_watch_hit", "制裁命中", "观察名单命中"],
    "transshipment_country": [
        "transshipment_country",
        "transit_country",
        "过境国",
        "转口国",
    ],
    "documents_complete_ratio": [
        "documents_complete_ratio",
        "docs_ratio",
        "单证完备度",
        "资料完整度",
    ],
    "kyc_verified": ["kyc_verified", "kyc", "KYC状态", "受益所有人核验"],
    "export_license_ready": [
        "export_license_ready",
        "license_ready",
        "许可证就绪",
        "出口许可",
    ],
    "production_buffer_days": ["production_buffer_days", "排产缓冲天数", "排产缓冲"],
    "logistics_delay_days": ["logistics_delay_days", "物流延误天数", "物流延误"],
    "gross_margin_ratio": ["gross_margin_ratio", "gross_margin", "毛利率"],
    "inventory_ready_ratio": [
        "inventory_ready_ratio",
        "inventory_ratio",
        "库存就绪度",
        "现货比例",
    ],
    "prior_dispute_count": ["prior_dispute_count", "历史争议次数", "争议次数"],
    "source_reputation": ["source_reputation", "线索可信度", "source_score"],
    "notes": ["notes", "备注", "note", "comment"],
    "raw_description": [
        "raw_description",
        "description",
        "other_info",
        "inquiry_text",
        "描述",
        "询盘描述",
        "需求描述",
    ],
}

PAYMENT_METHOD_PATTERNS = [
    (re.compile(r"insured\s*lc|保险信用证", re.IGNORECASE), "insured_lc"),
    (
        re.compile(r"letter\s*of\s*credit|\bl/c\b|\blc\b|信用证", re.IGNORECASE),
        "letter_of_credit",
    ),
    (re.compile(r"open\s*account|\boa\b|赊销|月结", re.IGNORECASE), "open_account"),
    (re.compile(r"tt\s*30\s*/\s*70|30/70|30%\s*deposit", re.IGNORECASE), "tt_30_70"),
    (
        re.compile(
            r"tt\s*after\s*bl|after\s*b/l|见提单后付款|放单后付款", re.IGNORECASE
        ),
        "tt_after_bl",
    ),
    (re.compile(r"\bda\b|承兑交单", re.IGNORECASE), "da"),
    (re.compile(r"dp\s*90|d/p\s*90|90天托收", re.IGNORECASE), "dp_90"),
]

INCOTERMS = ("EXW", "FCA", "FOB", "CFR", "CIF", "DAP", "DDP")

DUAL_USE_KEYWORDS = ("dual use", "dual-use", "军民两用", "servo", "伺服", "通信模组")
ELECTRONICS_KEYWORDS = ("pcb", "board", "control", "通信", "模组", "electronics")
MACHINERY_KEYWORDS = ("machine", "machinery", "夹具", "备件", "assembly", "装配")
CONSUMER_KEYWORDS = ("家电", "consumer", "套装", "home")

HIGH_SENSITIVITY_RULES = (
    {
        "name": "硬盘与电子废料",
        "keywords": (
            "hard disk scrap",
            "hdd scrap",
            "electronic scrap",
            "e-waste",
            "ewaste",
            "pcb scrap",
            "battery scrap",
        ),
        "product_category": "dual_use",
        "hs_code_sensitive": True,
        "export_license_ready": False,
    },
    {
        "name": "敏感电子部件",
        "keywords": (
            "capacitor",
            "capacitor cans",
            "servo",
            "communication module",
            "通信模组",
            "伺服",
        ),
        "product_category": None,
        "hs_code_sensitive": True,
        "export_license_ready": None,
    },
)

EXTERNAL_RISK_CATALOG_DIR = Path(__file__).resolve().parents[2] / "config"
EXTERNAL_RISK_CACHE_PATH = (
    Path(__file__).resolve().parents[2]
    / "data"
    / "cache"
    / "external_risk_runtime_cache.json"
)


def _load_external_risk_rules() -> tuple[dict[str, Any], ...]:
    rules: list[dict[str, Any]] = []
    for path in sorted(EXTERNAL_RISK_CATALOG_DIR.glob("external_risk_catalog*.json")):
        try:
            payload = json.loads(path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            continue
        for source in payload.get("sources", []) or []:
            for rule in source.get("rules", []) or []:
                merged = dict(rule)
                merged["source_id"] = source.get("source_id")
                merged["source_type"] = source.get("source_type")
                merged["provider"] = source.get("provider")
                merged["confidence"] = source.get("confidence")
                merged["priority"] = source.get("priority")
                merged["fallback_policy"] = source.get("fallback_policy")
                merged["selected_remote_rank"] = source.get("selected_remote_rank")
                merged["catalog_file"] = path.name
                rules.append(merged)
    if EXTERNAL_RISK_CACHE_PATH.exists():
        try:
            payload = json.loads(EXTERNAL_RISK_CACHE_PATH.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            payload = {}
        for source in payload.get("sources", []) or []:
            for rule in source.get("rules", []) or []:
                merged = dict(rule)
                merged["source_id"] = source.get("source_id")
                merged["source_type"] = source.get("source_type")
                merged["provider"] = source.get("provider")
                merged["confidence"] = source.get("confidence")
                merged["priority"] = source.get("priority")
                merged["fallback_policy"] = source.get("fallback_policy")
                merged["selected_remote_rank"] = source.get("selected_remote_rank")
                merged["catalog_file"] = str(
                    payload.get("cache_source") or EXTERNAL_RISK_CACHE_PATH.name
                )
                merged["cache_mode"] = "runtime_cache"
                rules.append(merged)
    return tuple(rules)


EXTERNAL_HIGH_SENSITIVITY_RULES = _load_external_risk_rules()


@dataclass(frozen=True)
class TradeFieldMappingRule:
    target_field: str
    source_aliases: tuple[str, ...]
    rule: str


FIELD_MAPPING_RULES = [
    TradeFieldMappingRule(
        target_field="payment_method",
        source_aliases=("payment_terms", "raw_description", "notes"),
        rule="优先读取付款条款列，再从原始描述中蒸馏 open account / lc / tt after bl / tt 30/70 等术语。",
    ),
    TradeFieldMappingRule(
        target_field="incoterm",
        source_aliases=("incoterm", "raw_description"),
        rule="从贸易术语列或描述中提取 EXW/FCA/FOB/CFR/CIF/DAP/DDP。",
    ),
    TradeFieldMappingRule(
        target_field="uses_free_email",
        source_aliases=("email", "uses_free_email"),
        rule="优先使用显式布尔列，否则根据邮箱域名是否属于免费邮箱集合判断。",
    ),
    TradeFieldMappingRule(
        target_field="product_category",
        source_aliases=("product_category", "product_name", "raw_description"),
        rule="优先使用显式类目，其次按双用途/电子/机械/消费品关键词推断。",
    ),
    TradeFieldMappingRule(
        target_field="risk_booleans",
        source_aliases=("raw_description", "notes", "asks_*"),
        rule="通过关键词匹配识别低报发票、第三方收款、先放单后付款、制裁命中与敏感 HS。",
    ),
]


class TradeInquiryAdapter:
    """将外贸询盘表转换为风控样本集。"""

    DEFAULT_AGENT_ROLES = [
        AgentRosterEntry(
            agent_id="counterparty_radar",
            agent_name="CounterpartyRadarAgent",
            focus="counterparty_payment",
        ),
        AgentRosterEntry(
            agent_id="compliance_sentinel",
            agent_name="ComplianceSentinelAgent",
            focus="compliance_sanctions",
        ),
        AgentRosterEntry(
            agent_id="fulfillment_pulse",
            agent_name="FulfillmentPulseAgent",
            focus="fulfillment_margin",
        ),
    ]

    def adapt_file(
        self,
        path: str | Path,
        *,
        sample_set_id: Optional[str] = None,
        cruise_goal: Optional[str] = None,
    ) -> TradeRiskSampleSet:
        source = Path(path)
        rows = self._read_rows(source)
        samples = [
            self._build_sample(index, row) for index, row in enumerate(rows, start=1)
        ]
        set_id = sample_set_id or f"trade_inquiry_import_{source.stem}"
        goal = (
            cruise_goal
            or "将外贸询盘表映射为 3-Agent 结构化风控样本，并接入巡航日报链路。"
        )
        return TradeRiskSampleSet(
            sample_set_id=set_id,
            cruise_goal=goal,
            agent_roles=list(self.DEFAULT_AGENT_ROLES),
            samples=samples,
        )

    @staticmethod
    def describe_mapping_rules() -> list[dict[str, Any]]:
        return [
            {
                "target_field": item.target_field,
                "source_aliases": list(item.source_aliases),
                "rule": item.rule,
            }
            for item in FIELD_MAPPING_RULES
        ]

    def _read_rows(self, path: Path) -> list[dict[str, Any]]:
        suffix = path.suffix.lower()
        if suffix == ".csv":
            with path.open("r", encoding="utf-8-sig", newline="") as handle:
                return [dict(row) for row in csv.DictReader(handle)]
        if suffix in {".xlsx", ".xlsm"}:
            try:
                from openpyxl import load_workbook
            except ImportError as exc:
                raise RuntimeError("读取 Excel 需要安装 openpyxl。") from exc
            workbook = load_workbook(path, read_only=True, data_only=True)
            sheet = workbook.active
            if sheet is None:
                return []
            sheet = cast(Any, sheet)
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                return []
            headers = [str(value or "").strip() for value in rows[0]]
            result = []
            for raw_row in rows[1:]:
                result.append(
                    {headers[idx]: raw_row[idx] for idx in range(len(headers))}
                )
            return result
        raise ValueError(f"不支持的文件格式: {path.suffix}")

    def _build_sample(self, index: int, row: dict[str, Any]) -> TradeRiskSample:
        canonical = self._canonicalize_row(row)
        notes = self._notes_from_row(canonical)
        requirement_text = str(canonical.get("requirement") or "").strip()
        frequency_text = str(canonical.get("frequency") or "").strip()
        payment_terms_text = str(canonical.get("payment_terms") or "").strip()
        raw_text = " ".join(
            str(canonical.get(key) or "")
            for key in (
                "raw_description",
                "notes",
                "payment_terms",
                "incoterm",
                "requirement",
                "frequency",
            )
        )
        email = str(canonical.get("email") or "").strip()
        product_name = self._infer_product_name(
            canonical.get("product_name"),
            canonical.get("scenario_name"),
            canonical.get("raw_description"),
            fallback=f"Imported Product {index}",
        )
        product_category = self._infer_product_category(
            canonical.get("product_category"), product_name, raw_text
        )
        payment_method = self._infer_payment_method(
            canonical.get("payment_terms"), raw_text
        )
        incoterm = self._infer_incoterm(canonical.get("incoterm"), raw_text)
        deposit_ratio = self._infer_deposit_ratio(
            canonical.get("deposit_ratio"), payment_method, payment_terms_text, raw_text
        )
        payment_days = self._infer_payment_days(
            canonical.get("payment_days"), payment_method, payment_terms_text, raw_text
        )
        uses_free_email = self._infer_free_email(
            canonical.get("uses_free_email"), email
        )
        hs_sensitive = (
            self._infer_bool(
                canonical.get("hs_code_sensitive"),
                raw_text,
                keywords=("敏感hs", "sensitive hs", "dual use"),
            )
            or product_category == "dual_use"
        )
        product_risk_profile = self._infer_product_risk_profile(product_name, raw_text)
        if product_risk_profile.get("product_category") == "dual_use":
            product_category = "dual_use"
        hs_sensitive = hs_sensitive or bool(
            product_risk_profile.get("hs_code_sensitive")
        )
        export_license_ready = self._infer_bool(
            canonical.get("export_license_ready"),
            raw_text,
            keywords=("license ready", "许可证齐全", "许可已备"),
            negative_keywords=("license pending", "许可证未办", "许可缺口"),
            default=not hs_sensitive,
        )
        if product_risk_profile.get("export_license_ready") is False:
            export_license_ready = False
        production_buffer_days = self._infer_production_buffer_days(
            canonical.get("production_buffer_days"),
            requirement_text,
            frequency_text,
            raw_text,
        )
        documents_complete_ratio = self._infer_documents_complete_ratio(
            canonical.get("documents_complete_ratio"),
            canonical,
            notes,
        )
        scenario_name = str(
            canonical.get("scenario_name") or ""
        ).strip() or self._build_scenario_name(
            canonical.get("destination_country"),
            product_name,
            payment_method,
            deposit_ratio,
        )
        return TradeRiskSample(
            sample_id=str(canonical.get("sample_id") or f"imported-trade-{index:03d}"),
            scenario_name=scenario_name,
            counterparty_name=str(
                canonical.get("counterparty_name") or f"Imported Buyer {index}"
            ),
            destination_country=str(canonical.get("destination_country") or "Unknown"),
            product_name=product_name,
            product_category=product_category,
            trade_value_usd=self._parse_money(
                canonical.get("trade_value_usd"), default=50000.0
            ),
            payment_method=payment_method,
            payment_days=payment_days,
            deposit_ratio=deposit_ratio,
            incoterm=incoterm,
            requested_discount_ratio=self._parse_ratio(
                canonical.get("requested_discount_ratio"),
                default=self._extract_percent(
                    raw_text, keywords=("discount", "折扣"), fallback=0.04
                ),
            ),
            account_age_days=self._parse_int(
                canonical.get("account_age_days"), default=180
            ),
            uses_free_email=uses_free_email,
            asks_undervalue_invoice=self._infer_bool(
                canonical.get("asks_undervalue_invoice"),
                raw_text,
                keywords=("低报", "undervalue", "lower invoice", "低开发票"),
                negative_keywords=(
                    "no undervalue",
                    "not undervalue",
                    "无低报",
                    "不低报",
                ),
            ),
            asks_split_payment_to_third_party=self._infer_bool(
                canonical.get("asks_split_payment_to_third_party"),
                raw_text,
                keywords=(
                    "第三方收款",
                    "third party payment",
                    "different beneficiary",
                    "保理代收",
                ),
                negative_keywords=(
                    "no third party payment",
                    "not third party",
                    "无第三方收款",
                ),
            ),
            asks_release_bl_before_payment=self._infer_bool(
                canonical.get("asks_release_bl_before_payment"),
                raw_text,
                keywords=("先放单", "release bl", "电放后付款", "after bl"),
                negative_keywords=("no release bl", "无需放单", "不要求放单"),
            ),
            hs_code_sensitive=hs_sensitive,
            sanction_watch_hit=self._infer_bool(
                canonical.get("sanction_watch_hit"),
                raw_text,
                keywords=("sanction hit", "制裁命中", "观察名单", "watchlist"),
                negative_keywords=(
                    "no sanction",
                    "no sanction hit",
                    "not sanctioned",
                    "未命中制裁",
                    "无制裁",
                ),
            ),
            transshipment_country=self._clean_optional_text(
                canonical.get("transshipment_country")
            ),
            documents_complete_ratio=documents_complete_ratio,
            kyc_verified=self._infer_bool(
                canonical.get("kyc_verified"),
                raw_text,
                keywords=("kyc ok", "kyc done", "已完成kyc", "ubo verified"),
                negative_keywords=(
                    "kyc pending",
                    "待做kyc",
                    "ubo missing",
                    "未完成kyc",
                ),
            ),
            export_license_ready=export_license_ready,
            production_buffer_days=production_buffer_days,
            logistics_delay_days=self._parse_int(
                canonical.get("logistics_delay_days"), default=4
            ),
            gross_margin_ratio=self._parse_ratio(
                canonical.get("gross_margin_ratio"),
                default=self._extract_percent(
                    raw_text, keywords=("margin", "毛利"), fallback=0.18
                ),
            ),
            inventory_ready_ratio=self._parse_ratio(
                canonical.get("inventory_ready_ratio"),
                default=self._extract_percent(
                    raw_text, keywords=("inventory", "库存"), fallback=0.85
                ),
            ),
            prior_dispute_count=self._parse_int(
                canonical.get("prior_dispute_count"), default=0
            ),
            source_reputation=self._parse_ratio(
                canonical.get("source_reputation"),
                default=self._infer_source_reputation(raw_text),
            ),
            external_match_count=int(
                product_risk_profile.get("external_match_count") or 0
            ),
            external_conflict_level=str(
                product_risk_profile.get("external_conflict_level") or "none"
            ),
            external_match_sources=[
                str(item)
                for item in product_risk_profile.get("external_match_sources", [])
                if str(item or "").strip()
            ],
            external_resolution_advice=str(
                product_risk_profile.get("external_resolution_advice") or "no_action"
            ),
            external_conflict_summary=dict(
                product_risk_profile.get("external_conflict_summary") or {}
            ),
            notes=notes,
        )

    def _canonicalize_row(self, row: dict[str, Any]) -> dict[str, Any]:
        normalized = {self._normalize_header(key): value for key, value in row.items()}
        canonical: dict[str, Any] = {}
        for target, aliases in HEADER_ALIASES.items():
            for alias in aliases:
                value = normalized.get(self._normalize_header(alias))
                if value not in (None, ""):
                    canonical[target] = value
                    break
        return canonical

    @staticmethod
    def _normalize_header(value: Any) -> str:
        return re.sub(r"[^a-z0-9\u4e00-\u9fff]+", "", str(value or "").strip().lower())

    @staticmethod
    def _clean_optional_text(value: Any) -> Optional[str]:
        text = str(value or "").strip()
        return text or None

    @staticmethod
    def _build_scenario_name(
        country: Any, product_name: str, payment_method: str, deposit_ratio: float
    ) -> str:
        destination = str(country or "未知目的国")
        return f"{destination}询盘：{product_name} | {payment_method} | 首付{int(round(deposit_ratio * 100))}%"

    @staticmethod
    def _infer_product_name(
        explicit: Any, scenario_name: Any, raw_description: Any, *, fallback: str
    ) -> str:
        candidates = [explicit, scenario_name, raw_description]
        for candidate in candidates:
            text = str(candidate or "").strip()
            if not text:
                continue
            lowered = text.lower()
            prefixes = (
                "looking for ",
                "i want to buy ",
                "i am interested in buying ",
            )
            for prefix in prefixes:
                if lowered.startswith(prefix):
                    text = text[len(prefix) :].strip(" .:-")
                    break
            if text:
                return text
        return fallback

    @staticmethod
    def _notes_from_row(canonical: dict[str, Any]) -> list[str]:
        notes: list[str] = []
        raw_notes = str(canonical.get("notes") or "").strip()
        if raw_notes:
            notes.extend(
                [
                    item.strip()
                    for item in re.split(r"[;；\n]+", raw_notes)
                    if item.strip()
                ]
            )
        raw_description = str(canonical.get("raw_description") or "").strip()
        if raw_description:
            notes.append(f"raw_description={raw_description[:220]}")
        return notes[:6]

    @staticmethod
    def _parse_int(value: Any, *, default: int) -> int:
        if value in (None, ""):
            return default
        text = str(value).strip()
        match = re.search(r"-?\d+", text.replace(",", ""))
        return int(match.group()) if match else default

    @staticmethod
    def _parse_money(value: Any, *, default: float) -> float:
        if value in (None, ""):
            return default
        text = str(value).strip().replace(",", "")
        match = re.search(r"-?\d+(?:\.\d+)?", text)
        return float(match.group()) if match else default

    @staticmethod
    def _parse_ratio(value: Any, *, default: float) -> float:
        if value in (None, ""):
            return default
        text = str(value).strip()
        match = re.search(r"-?\d+(?:\.\d+)?", text.replace(",", ""))
        if not match:
            return default
        number = float(match.group())
        if "%" in text or number > 1:
            number = number / 100.0
        return max(0.0, min(1.0, number))

    def _infer_payment_method(self, explicit: Any, raw_text: str) -> str:
        text = f"{explicit or ''} {raw_text}".strip()
        lowered = text.lower()
        if any(
            token in lowered
            for token in (
                "online",
                "as per suppliers",
                "as per supplier",
                "upi",
                "bank transfer",
            )
        ):
            return "tt_30_70"
        for pattern, label in PAYMENT_METHOD_PATTERNS:
            if pattern.search(text):
                return label
        # 对真实询盘标题缺乏付款术语时，默认按更保守的赊销场景处理，避免漏报。
        return "open_account"

    def _infer_payment_days(
        self, explicit: Any, payment_method: str, payment_terms: str, raw_text: str
    ) -> int:
        parsed = self._parse_int(explicit, default=-1)
        if parsed >= 0:
            return parsed
        lowered = f"{payment_terms} {raw_text}".lower()
        if any(token in lowered for token in ("after bl", "after b/l", "release bl")):
            return 90
        if payment_method in {"tt_after_bl", "dp_90"}:
            return 60
        if payment_method in {"open_account", "da"}:
            return 45
        if payment_method == "letter_of_credit":
            return 0
        return 30

    def _infer_deposit_ratio(
        self, explicit: Any, payment_method: str, payment_terms: str, raw_text: str
    ) -> float:
        parsed = self._parse_ratio(explicit, default=-1.0)
        if parsed >= 0:
            return parsed
        lowered = f"{payment_terms} {raw_text}".lower()
        if any(
            token in lowered
            for token in ("100% advance", "full advance", "advance payment")
        ):
            return 1.0
        if payment_method in {"open_account", "tt_after_bl", "da", "dp_90"}:
            # 未给出明确首付时，赊销类条款按更保守的 15% 安全垫估计。
            return 0.15
        if payment_method == "tt_30_70":
            return 0.3
        if payment_method == "letter_of_credit":
            return 0.2
        return 0.3

    def _infer_production_buffer_days(
        self, explicit: Any, requirement_text: str, frequency_text: str, raw_text: str
    ) -> int:
        parsed = self._parse_int(explicit, default=-1)
        if parsed >= 0:
            return parsed
        lowered = f"{requirement_text} {frequency_text} {raw_text}".lower()
        if any(
            token in lowered
            for token in (
                "urgent",
                "immediately",
                "immediate",
                "asap",
                "within 3 days",
                "within 5 days",
            )
        ):
            return 0
        if any(
            token in lowered
            for token in ("within 1 week", "within one week", "7 days", "一周内")
        ):
            return 2
        if any(
            token in lowered
            for token in ("within 2 weeks", "within two weeks", "2 weeks", "两周内")
        ):
            return 8
        if "monthly" in lowered:
            return 10
        return 12

    def _infer_documents_complete_ratio(
        self,
        explicit: Any,
        canonical: dict[str, Any],
        notes: list[str],
    ) -> float:
        parsed = self._parse_ratio(explicit, default=-1.0)
        if parsed >= 0:
            return parsed
        evidence_count = 0
        for key in (
            "counterparty_name",
            "destination_country",
            "quantity",
            "requirement",
            "frequency",
            "payment_terms",
            "raw_description",
        ):
            if str(canonical.get(key) or "").strip():
                evidence_count += 1
        if evidence_count >= 6:
            return 0.84
        if evidence_count >= 4:
            return 0.78
        if notes:
            return 0.62
        return 0.58

    @staticmethod
    def _infer_incoterm(explicit: Any, raw_text: str) -> str:
        text = f"{explicit or ''} {raw_text}".upper()
        for term in INCOTERMS:
            if re.search(rf"\b{term}\b", text):
                return term
        return "FOB"

    @staticmethod
    def _infer_product_category(explicit: Any, product_name: str, raw_text: str) -> str:
        if explicit:
            return str(explicit).strip().lower()
        text = f"{product_name} {raw_text}".lower()
        if any(keyword in text for keyword in DUAL_USE_KEYWORDS):
            return "dual_use"
        if any(keyword in text for keyword in ELECTRONICS_KEYWORDS):
            return "electronics"
        if any(keyword in text for keyword in MACHINERY_KEYWORDS):
            return "machinery"
        if any(keyword in text for keyword in CONSUMER_KEYWORDS):
            return "consumer_goods"
        return "electronics"

    @staticmethod
    def _infer_product_risk_profile(product_name: str, raw_text: str) -> dict[str, Any]:
        text = f"{product_name} {raw_text}".lower()
        external_matches = []
        for rule in EXTERNAL_HIGH_SENSITIVITY_RULES:
            if any(keyword in text for keyword in rule.get("keywords", [])):
                external_matches.append(rule)
        if external_matches:
            first = external_matches[0]
            categories = {
                str(item.get("product_category"))
                for item in external_matches
                if item.get("product_category") is not None
            }
            conflict_level = "aligned" if len(categories) <= 1 else "conflict"
            conflict_summary = TradeInquiryAdapter._build_external_conflict_summary(
                external_matches, conflict_level=conflict_level
            )
            return {
                "rule_name": first.get("rule_id")
                or first.get("source_id")
                or "external_rule",
                "product_category": first.get("product_category"),
                "hs_code_sensitive": first.get("hs_code_sensitive"),
                "export_license_ready": first.get("export_license_ready"),
                "external_source": first.get("source_id"),
                "external_provider": first.get("provider"),
                "risk_note": first.get("risk_note"),
                "external_match_count": len(external_matches),
                "external_conflict_level": conflict_level,
                "external_match_sources": [
                    item.get("source_id") for item in external_matches
                ],
                "external_resolution_advice": (
                    "accept_aligned" if conflict_level == "aligned" else "manual_review"
                ),
                "external_conflict_summary": conflict_summary,
            }
        for rule in HIGH_SENSITIVITY_RULES:
            if any(keyword in text for keyword in rule["keywords"]):
                return {
                    "rule_name": rule["name"],
                    "product_category": rule["product_category"],
                    "hs_code_sensitive": rule["hs_code_sensitive"],
                    "export_license_ready": rule["export_license_ready"],
                }
        return {
            "rule_name": "",
            "product_category": None,
            "hs_code_sensitive": False,
            "export_license_ready": None,
            "external_source": None,
            "external_provider": None,
            "risk_note": None,
            "external_match_count": 0,
            "external_conflict_level": "none",
            "external_match_sources": [],
            "external_resolution_advice": "no_action",
            "external_conflict_summary": {},
        }

    @staticmethod
    def _build_external_conflict_summary(
        external_matches: list[dict[str, Any]], *, conflict_level: str
    ) -> dict[str, Any]:
        by_source: dict[str, dict[str, Any]] = {}
        for item in external_matches:
            source_id = str(item.get("source_id") or "unknown")
            entry = by_source.setdefault(
                source_id,
                {
                    "source_id": source_id,
                    "priority": int(item.get("priority") or 999),
                    "confidence": float(item.get("confidence") or 0.5),
                    "match_count": 0,
                    "rule_ids": [],
                    "selected_remote_rank": item.get("selected_remote_rank"),
                    "fallback_policy": item.get("fallback_policy"),
                },
            )
            entry["match_count"] += 1
            rule_id = str(item.get("rule_id") or "unknown_rule")
            if rule_id not in entry["rule_ids"]:
                entry["rule_ids"].append(rule_id)
        ordered = sorted(
            by_source.values(),
            key=lambda item: (
                int(item.get("priority") or 999),
                -float(item.get("confidence") or 0.0),
                -int(item.get("match_count") or 0),
            ),
        )
        top = ordered[0] if ordered else None
        return {
            "ordered_sources": [item["source_id"] for item in ordered],
            "top_source": top.get("source_id") if top else None,
            "top_source_priority": top.get("priority") if top else None,
            "top_source_confidence": top.get("confidence") if top else None,
            "top_source_rule_ids": top.get("rule_ids", []) if top else [],
            "source_breakdown": ordered,
            "resolution_basis": (
                "prefer_highest_priority_source"
                if conflict_level == "aligned"
                else "manual_review_with_priority_context"
            ),
        }

    def _infer_free_email(self, explicit: Any, email: str) -> bool:
        parsed = self._parse_bool(explicit)
        if parsed is not None:
            return parsed
        domain = email.split("@")[-1].lower() if "@" in email else ""
        return domain in FREE_EMAIL_DOMAINS

    def _infer_bool(
        self,
        explicit: Any,
        raw_text: str,
        *,
        keywords: Iterable[str],
        negative_keywords: Iterable[str] = (),
        default: bool = False,
    ) -> bool:
        parsed = self._parse_bool(explicit)
        if parsed is not None:
            return parsed
        lowered = raw_text.lower()
        if any(keyword.lower() in lowered for keyword in negative_keywords):
            return False
        if any(keyword.lower() in lowered for keyword in keywords):
            return True
        return default

    @staticmethod
    def _parse_bool(value: Any) -> Optional[bool]:
        if value in (None, ""):
            return None
        lowered = str(value).strip().lower()
        if lowered in {"1", "true", "yes", "y", "是", "有", "需要", "命中", "已完成"}:
            return True
        if lowered in {
            "0",
            "false",
            "no",
            "n",
            "否",
            "无",
            "不需要",
            "未命中",
            "未完成",
        }:
            return False
        return None

    @staticmethod
    def _extract_percent(
        raw_text: str, *, keywords: Iterable[str], fallback: float
    ) -> float:
        lowered = raw_text.lower()
        if not any(keyword.lower() in lowered for keyword in keywords):
            return fallback
        match = re.search(r"(\d+(?:\.\d+)?)\s*%", raw_text)
        if not match:
            return fallback
        return max(0.0, min(1.0, float(match.group(1)) / 100.0))

    @staticmethod
    def _infer_source_reputation(raw_text: str) -> float:
        lowered = raw_text.lower()
        if any(
            token in lowered
            for token in ("linkedin", "customs", "海关", "mofcom", "gov")
        ):
            return 0.82
        if any(
            token in lowered for token in ("whatsapp", "unknown", "转介绍", "broker")
        ):
            return 0.58
        return 0.72
